from __future__ import annotations

import io
import logging
import os
import threading
import uuid
import zipfile
from datetime import datetime, timedelta, timezone

import requests as requests_lib
from dotenv import load_dotenv
from fastapi import BackgroundTasks, Body, Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db
from .auth import get_current_user, require_admin, hash_password, verify_password, create_token
from .manual_finder import find_manual_and_warranty, download_pdf_from_url
from .pdf_parser import parse_products_from_pdf

# Load environment variables early
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("ati")

app = FastAPI(title="ATI Manual Finder")

# ---------------------------------------------------------------------------
# In-memory progress tracker
# ---------------------------------------------------------------------------
_progress: dict[str, dict] = {}
_download_jobs: dict[str, dict] = {}


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

def _process_project(project_id: str, total: int):
    """Process all pending items for a project (runs in background thread)."""
    try:
        items = db.get_project_items(project_id)
        pending = [i for i in items if i["status"] == "pending"]
        logger.info("Starting processing for project %s — %d items", project_id, len(pending))

        for idx, item in enumerate(pending, 1):
            brand = item.get("brand", "")
            model = item["model_number"]
            name = item.get("product_name", "")

            _progress[project_id] = {
                "message": f"Searching {idx}/{total}: {brand} {model}",
                "done": False,
                "error": None,
            }

            try:
                # Check product library cache first
                cached = db.get_product_by_model(model)
                if cached:
                    link_data = {"product_id": cached["id"]}

                    if cached.get("manual_source_url"):
                        # FULL CACHE HIT — manual exists in library
                        logger.info("[%d/%d] Cache hit (full): %s %s", idx, total, brand, model)
                        manual_url = None
                        if cached.get("manual_storage_path"):
                            try:
                                manual_url = db.get_manual_url(cached["manual_storage_path"])
                            except Exception:
                                manual_url = cached.get("manual_source_url")
                        link_data["status"] = "found"
                        link_data["manual_url"] = manual_url or cached.get("manual_source_url")
                        db.update_project_item(item["id"], link_data)
                        continue

                    # Check if recently verified as not_found (within 7 days)
                    last_verified = cached.get("last_verified")
                    if last_verified:
                        try:
                            lv = datetime.fromisoformat(str(last_verified).replace("Z", "+00:00"))
                            if lv > datetime.now(timezone.utc) - timedelta(days=7):
                                logger.info("[%d/%d] Cache hit (recent not_found): %s %s", idx, total, brand, model)
                                link_data["status"] = "not_found"
                                link_data["notes"] = "Previously searched (cached)"
                                db.update_project_item(item["id"], link_data)
                                continue
                        except Exception:
                            pass

                    # Partial cache — re-search but carry over warranty
                    logger.info("[%d/%d] Partial cache (re-searching): %s %s", idx, total, brand, model)
                    db.update_project_item(item["id"], link_data)

                # Web search
                logger.info("[%d/%d] Searching web: %s %s", idx, total, brand, model)
                existing_warranty = cached.get("warranty_length") if cached else None
                result = find_manual_and_warranty(brand, model, name)
                if existing_warranty and not result.get("warranty_length"):
                    result["warranty_length"] = existing_warranty
                logger.info(
                    "[%d/%d] Result: status=%s, manual_url=%s, warranty=%s",
                    idx, total, result["status"],
                    result.get("manual_source_url", "none"),
                    result.get("warranty_length", "none"),
                )

                storage_path = None
                manual_url = None

                if result["manual_pdf_bytes"]:
                    safe_brand = (brand or "unknown").replace(" ", "_")
                    storage_path = f"{project_id}/{safe_brand}_{model}_manual.pdf"
                    try:
                        db.upload_manual(result["manual_pdf_bytes"], storage_path)
                        manual_url = db.get_manual_url(storage_path)
                        logger.info("[%d/%d] Uploaded manual to storage: %s", idx, total, storage_path)
                    except Exception as e:
                        logger.warning("[%d/%d] Storage upload failed: %s", idx, total, e)
                        manual_url = result.get("manual_source_url")

                # Upsert product library — prefer Supabase signed URL over original website URL
                product = db.upsert_product({
                    "brand": brand,
                    "model_number": model,
                    "product_name": name,
                    "manual_source_url": manual_url or result.get("manual_source_url"),
                    "manual_storage_path": storage_path,
                    "warranty_length": result.get("warranty_length"),
                    "last_verified": datetime.now(timezone.utc).isoformat(),
                })

                # Update project item
                update_data = {
                    "status": result["status"],
                    "manual_url": manual_url or result.get("manual_source_url"),
                }
                if product.get("id"):
                    update_data["product_id"] = product["id"]
                db.update_project_item(item["id"], update_data)

            except Exception as e:
                logger.error("[%d/%d] Failed to process %s %s: %s", idx, total, brand, model, e, exc_info=True)
                try:
                    db.update_project_item(item["id"], {
                        "status": "not_found",
                        "notes": f"Error: {e}",
                    })
                except Exception:
                    pass

        _progress[project_id] = {"message": "Complete!", "done": True, "error": None}
        logger.info("Processing complete for project %s", project_id)

    except Exception as e:
        logger.error("Worker crashed for project %s: %s", project_id, e, exc_info=True)
        _progress[project_id] = {"message": str(e), "done": True, "error": str(e)}


# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------

@app.get("/api/health")
async def health():
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Auth routes (public — no token required)
# ---------------------------------------------------------------------------

@app.post("/api/auth/register")
async def register(data: dict):
    email = (data.get("email") or "").strip().lower()
    password = data.get("password", "")
    name = (data.get("name") or "").strip()

    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Valid email is required")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    existing = db.get_user_by_email(email)
    if existing:
        raise HTTPException(status_code=409, detail="Email already registered")

    # First user auto-becomes admin and auto-approved
    user_count = db.count_users()
    if user_count == 0:
        role = "admin"
        approved = True
        message = "Admin account created. You can log in now."
    else:
        role = "pm"
        approved = False
        message = "Account created. Waiting for admin approval."

    password_hash = hash_password(password)
    user = db.create_user(email=email, password_hash=password_hash, name=name, role=role, approved=approved)

    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "approved": user["approved"],
        "message": message,
    }


@app.post("/api/auth/login")
async def login(data: dict):
    email = (data.get("email") or "").strip().lower()
    password = data.get("password", "")

    user = db.get_user_by_email(email)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not verify_password(password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not user.get("approved"):
        raise HTTPException(status_code=403, detail="Account pending admin approval")

    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
        },
    }


@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "approved": user.get("approved"),
    }


# ---------------------------------------------------------------------------
# Admin user management (require admin role)
# ---------------------------------------------------------------------------

@app.get("/api/admin/users")
async def admin_list_users(user: dict = Depends(require_admin)):
    return db.list_users()


@app.patch("/api/admin/users/{user_id}")
async def admin_update_user(user_id: str, data: dict, user: dict = Depends(require_admin)):
    # Safety: cannot change own role or unapprove yourself
    if user_id == user["id"]:
        if "role" in data and data["role"] != user["role"]:
            raise HTTPException(status_code=400, detail="Cannot change your own role")
        if "approved" in data and not data["approved"]:
            raise HTTPException(status_code=400, detail="Cannot unapprove yourself")
    return db.update_user(user_id, data)


@app.delete("/api/admin/users/{user_id}")
async def admin_delete_user(user_id: str, user: dict = Depends(require_admin)):
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    db.delete_user(user_id)
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# Helper: project ownership check
# ---------------------------------------------------------------------------

async def _check_project_access(project_id: str, user: dict):
    """Raise 403 if PM doesn't own this project. Admin can access all."""
    if user["role"] == "admin":
        return
    project = db.get_project(project_id)
    if not project or project.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")


# ---------------------------------------------------------------------------
# Project routes (protected)
# ---------------------------------------------------------------------------

@app.get("/api/projects")
async def list_projects(user: dict = Depends(get_current_user)):
    if user["role"] == "admin":
        return db.list_projects()  # all projects with user info
    return db.list_projects(user_id=user["id"])


@app.post("/api/projects/upload")
async def upload_project(
    background_tasks: BackgroundTasks,
    project_name: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="Admin accounts cannot create projects")

    pdf_bytes = await file.read()
    logger.info("Uploading project '%s', PDF size: %d bytes", project_name, len(pdf_bytes))

    # Parse products from PDF
    products = parse_products_from_pdf(pdf_bytes)
    logger.info("Extracted %d products from PDF", len(products))

    # Create project with user_id
    project = db.create_project(name=project_name, user_id=user["id"])
    project_id = project["id"]

    # Create project items
    items = [
        {
            "project_id": project_id,
            "brand": p.get("brand"),
            "model_number": p["model_number"],
            "product_name": p.get("product_name"),
            "raw_line_item": f"{p.get('brand', '')} {p['model_number']}",
            "status": "pending",
        }
        for p in products
    ]
    if items:
        db.create_project_items(items)

    # Initialize progress
    _progress[project_id] = {"message": "Starting...", "done": False, "error": None}

    # Kick off background processing
    background_tasks.add_task(_process_project, project_id, len(products))

    return {
        "project_id": project_id,
        "product_count": len(products),
        "products": products,
    }


@app.get("/api/projects/{project_id}")
async def get_project(project_id: str, user: dict = Depends(get_current_user)):
    await _check_project_access(project_id, user)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    items = db.get_project_items(project_id)
    progress = _progress.get(project_id, {"message": "Not started", "done": True, "error": None})
    return {"project": project, "items": items, "progress": progress}


@app.post("/api/projects/{project_id}/add-items")
async def add_items_to_project(
    project_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Upload an additional PDF to an existing project (change orders)."""
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="Admin accounts cannot add items to projects")
    await _check_project_access(project_id, user)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    pdf_bytes = await file.read()
    logger.info("Adding items to project '%s', PDF size: %d bytes", project["name"], len(pdf_bytes))

    # Parse products from PDF
    products = parse_products_from_pdf(pdf_bytes)
    logger.info("Extracted %d products from additional PDF", len(products))

    if not products:
        return {"project_id": project_id, "product_count": 0, "products": []}

    # Deduplicate against existing items in this project
    existing_items = db.get_project_items(project_id)
    existing_models = {i["model_number"].upper() for i in existing_items if i.get("model_number")}

    new_products = [p for p in products if p["model_number"].upper() not in existing_models]
    logger.info("After dedup: %d new products (skipped %d duplicates)", len(new_products), len(products) - len(new_products))

    if not new_products:
        return {"project_id": project_id, "product_count": 0, "products": [], "message": "All products already exist in this project"}

    # Create new project items
    items = [
        {
            "project_id": project_id,
            "brand": p.get("brand"),
            "model_number": p["model_number"],
            "product_name": p.get("product_name"),
            "raw_line_item": f"{p.get('brand', '')} {p['model_number']}",
            "status": "pending",
        }
        for p in new_products
    ]
    db.create_project_items(items)

    # Initialize/update progress
    _progress[project_id] = {"message": "Starting...", "done": False, "error": None}

    # Kick off background processing for new items only
    background_tasks.add_task(_process_project, project_id, len(new_products))

    return {
        "project_id": project_id,
        "product_count": len(new_products),
        "products": new_products,
        "skipped": len(products) - len(new_products),
    }


@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: str, user: dict = Depends(get_current_user)):
    await _check_project_access(project_id, user)
    db.delete_project(project_id)
    _progress.pop(project_id, None)
    return {"status": "deleted"}


@app.get("/api/projects/{project_id}/progress")
async def get_progress(project_id: str, user: dict = Depends(get_current_user)):
    await _check_project_access(project_id, user)
    return _progress.get(
        project_id, {"message": "Not started", "done": True, "error": None}
    )


def _download_manual_from_url(
    item_id: str, url: str, brand: str, model: str, project_id: str
):
    """Background task: download a PDF from a user-pasted URL and upload to Supabase Storage."""
    try:
        pdf_bytes = download_pdf_from_url(url)
        if not pdf_bytes:
            logger.info("Background download: no PDF obtained from %s", url[:80])
            return

        safe_brand = (brand or "unknown").replace(" ", "_")
        storage_path = f"{project_id}/{safe_brand}_{model}_manual.pdf"
        db.upload_manual(pdf_bytes, storage_path)
        signed_url = db.get_manual_url(storage_path)

        logger.info("Background download: uploaded to %s", storage_path)

        # Update product cache with storage path
        db.upsert_product({
            "model_number": model,
            "brand": brand,
            "manual_storage_path": storage_path,
            "manual_source_url": signed_url,
            "last_verified": datetime.now(timezone.utc).isoformat(),
        })

        # Update project item to use the signed URL
        db.update_project_item(item_id, {"manual_url": signed_url})

    except Exception as e:
        logger.warning("Background download failed for %s: %s", url[:80], e)


@app.patch("/api/items/{item_id}")
async def update_item(item_id: str, data: dict, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    # Extract warranty_length — it belongs on the products table, not project_items
    warranty_length = data.pop("warranty_length", None)

    # If manual_url is provided, set status to manual_entry
    if "manual_url" in data and data["manual_url"]:
        data.setdefault("status", "manual_entry")

    # Update the project item
    updated_item = db.update_project_item(item_id, data) if data else {}

    # Also update the products table (cache) so future projects get a cache hit
    manual_url = data.get("manual_url")
    if manual_url or warranty_length:
        sb = db.get_client()
        resp = sb.table("project_items").select("*").eq("id", item_id).maybe_single().execute()
        if resp and resp.data:
            item = resp.data
            model_number = item.get("model_number")
            if model_number:
                product_data = {
                    "model_number": model_number,
                    "brand": item.get("brand"),
                    "product_name": item.get("product_name"),
                }
                if manual_url:
                    product_data["manual_source_url"] = manual_url
                if warranty_length:
                    product_data["warranty_length"] = warranty_length
                product_data["last_verified"] = datetime.now(timezone.utc).isoformat()

                product = db.upsert_product(product_data)

                # Link the project item to the product
                if product.get("id"):
                    db.update_project_item(item_id, {"product_id": product["id"]})

            # Trigger background PDF download from pasted URL
            if manual_url:
                background_tasks.add_task(
                    _download_manual_from_url,
                    item_id,
                    manual_url,
                    item.get("brand", ""),
                    item.get("model_number", ""),
                    item.get("project_id", ""),
                )

    # Return item with product join so frontend gets fresh warranty
    sb = db.get_client()
    final = sb.table("project_items").select("*, products(*)").eq("id", item_id).maybe_single().execute()
    return final.data if final and final.data else updated_item


@app.post("/api/items/{item_id}/retry")
async def retry_item(item_id: str, user: dict = Depends(get_current_user)):
    # Fetch the item to get brand/model/product_name
    sb = db.get_client()
    resp = sb.table("project_items").select("*").eq("id", item_id).maybe_single().execute()
    if resp is None or not resp.data:
        raise HTTPException(status_code=404, detail="Item not found")
    item = resp.data

    logger.info("Retrying item: %s %s", item.get("brand", ""), item["model_number"])

    result = find_manual_and_warranty(
        item.get("brand", ""),
        item["model_number"],
        item.get("product_name", ""),
    )

    storage_path = None
    manual_url = None

    if result["manual_pdf_bytes"]:
        safe_brand = (item.get("brand") or "unknown").replace(" ", "_")
        storage_path = f"{item['project_id']}/{safe_brand}_{item['model_number']}_manual.pdf"
        try:
            db.upload_manual(result["manual_pdf_bytes"], storage_path)
            manual_url = db.get_manual_url(storage_path)
        except Exception:
            manual_url = result.get("manual_source_url")

    product = db.upsert_product({
        "brand": item.get("brand"),
        "model_number": item["model_number"],
        "product_name": item.get("product_name"),
        "manual_source_url": manual_url or result.get("manual_source_url"),
        "manual_storage_path": storage_path,
        "warranty_length": result.get("warranty_length"),
        "last_verified": datetime.now(timezone.utc).isoformat(),
    })

    update_data = {
        "status": result["status"],
        "manual_url": manual_url or result.get("manual_source_url"),
    }
    if product.get("id"):
        update_data["product_id"] = product["id"]

    updated = db.update_project_item(item_id, update_data)
    return updated


@app.delete("/api/items/{item_id}")
async def delete_item(item_id: str, user: dict = Depends(get_current_user)):
    db.delete_project_item(item_id)
    return {"status": "deleted"}


@app.post("/api/items/batch-delete")
async def batch_delete_items(data: dict, user: dict = Depends(get_current_user)):
    item_ids = data.get("item_ids", [])
    if not item_ids:
        raise HTTPException(status_code=400, detail="item_ids required")
    db.delete_project_items(item_ids)
    return {"status": "deleted", "count": len(item_ids)}


# ---------------------------------------------------------------------------
# Product library admin
# ---------------------------------------------------------------------------

@app.get("/api/products")
async def list_products(search: str | None = None, user: dict = Depends(get_current_user)):
    products = db.list_products(search)
    # Regenerate fresh signed URLs for products stored in Supabase
    for p in products:
        if p.get("manual_storage_path"):
            try:
                p["manual_source_url"] = db.get_manual_url(p["manual_storage_path"])
            except Exception:
                pass  # keep existing manual_source_url if signing fails
    return products


@app.post("/api/products")
async def create_product(data: dict, user: dict = Depends(get_current_user)):
    if not data.get("model_number"):
        raise HTTPException(status_code=400, detail="model_number is required")
    return db.upsert_product(data)


@app.patch("/api/products/{product_id}")
async def update_product(product_id: str, data: dict, user: dict = Depends(get_current_user)):
    return db.update_product(product_id, data)


@app.delete("/api/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(get_current_user)):
    db.delete_product(product_id)
    return {"status": "deleted"}


@app.post("/api/products/{product_id}/upload-manual")
async def upload_product_manual(product_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a PDF manual for a product in the library."""
    sb = db.get_client()
    resp = sb.table("products").select("*").eq("id", product_id).maybe_single().execute()
    if not resp or not resp.data:
        raise HTTPException(status_code=404, detail="Product not found")
    product = resp.data

    pdf_bytes = await file.read()
    safe_brand = (product.get("brand") or "unknown").replace(" ", "_")
    model = product["model_number"]
    storage_path = f"library/{safe_brand}_{model}_manual.pdf"

    db.upload_manual(pdf_bytes, storage_path)
    manual_url = db.get_manual_url(storage_path)

    updated = db.update_product(product_id, {
        "manual_storage_path": storage_path,
        "manual_source_url": manual_url,
        "last_verified": datetime.now(timezone.utc).isoformat(),
    })
    return updated


@app.post("/api/items/{item_id}/upload-manual")
async def upload_item_manual(item_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a PDF manual for a project item (also updates product cache)."""
    sb = db.get_client()
    resp = sb.table("project_items").select("*").eq("id", item_id).maybe_single().execute()
    if not resp or not resp.data:
        raise HTTPException(status_code=404, detail="Item not found")
    item = resp.data

    pdf_bytes = await file.read()
    safe_brand = (item.get("brand") or "unknown").replace(" ", "_")
    model = item["model_number"]
    storage_path = f"{item['project_id']}/{safe_brand}_{model}_manual.pdf"

    db.upload_manual(pdf_bytes, storage_path)
    manual_url = db.get_manual_url(storage_path)

    # Update product cache
    product = db.upsert_product({
        "brand": item.get("brand"),
        "model_number": model,
        "product_name": item.get("product_name"),
        "manual_source_url": manual_url,
        "manual_storage_path": storage_path,
        "last_verified": datetime.now(timezone.utc).isoformat(),
    })

    # Update project item
    update_data = {
        "status": "manual_entry",
        "manual_url": manual_url,
    }
    if product.get("id"):
        update_data["product_id"] = product["id"]
    db.update_project_item(item_id, update_data)

    # Return with product join
    final = sb.table("project_items").select("*, products(*)").eq("id", item_id).maybe_single().execute()
    return final.data if final and final.data else update_data


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@app.get("/api/projects/{project_id}/export")
async def export_excel(project_id: str, ids: str | None = None, user: dict = Depends(get_current_user)):
    await _check_project_access(project_id, user)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    all_items = db.get_project_items(project_id)
    if ids:
        selected = set(ids.split(","))
        items = [i for i in all_items if i["id"] in selected]
    else:
        items = all_items

    wb = Workbook()

    # --- Sheet 1: Products & Manuals ---
    ws1 = wb.active
    ws1.title = "Products & Manuals"

    navy_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")

    headers = ["Brand", "Model Number", "Product Name", "Warranty", "Manual Link", "Status", "Notes"]
    widths = [15, 22, 45, 20, 20, 15, 30]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_bold
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    status_colors = {
        "found": "C6EFCE",
        "not_found": "FFC7CE",
        "manual_entry": "FFEB9C",
        "pending": "D9D9D9",
    }

    for row_idx, item in enumerate(items, 2):
        ws1.cell(row=row_idx, column=1, value=item.get("brand"))
        ws1.cell(row=row_idx, column=2, value=item.get("model_number"))
        ws1.cell(row=row_idx, column=3, value=item.get("product_name"))

        # Warranty from linked product
        product = item.get("products")
        warranty = product.get("warranty_length") if product else None
        ws1.cell(row=row_idx, column=4, value=warranty)

        # Manual link as hyperlink
        manual_url = item.get("manual_url")
        if manual_url:
            cell = ws1.cell(row=row_idx, column=5, value="Open Manual")
            cell.hyperlink = manual_url
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws1.cell(row=row_idx, column=5, value="")

        # Status with color
        status = item.get("status", "pending")
        status_cell = ws1.cell(row=row_idx, column=6, value=status)
        if status in status_colors:
            status_cell.fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type="solid",
            )

        ws1.cell(row=row_idx, column=7, value=item.get("notes"))

    # --- Sheet 2: Needs Manual Lookup ---
    ws2 = wb.create_sheet("Needs Manual Lookup")
    not_found = [i for i in items if i.get("status") == "not_found"]

    headers2 = [
        "Brand", "Model Number", "Product Name",
        "Manual URL (paste here)", "Warranty (enter here)", "Notes",
    ]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_bold

    for row_idx, item in enumerate(not_found, 2):
        ws2.cell(row=row_idx, column=1, value=item.get("brand"))
        ws2.cell(row=row_idx, column=2, value=item.get("model_number"))
        ws2.cell(row=row_idx, column=3, value=item.get("product_name"))
        # Columns 4, 5, 6 left blank for user to fill in

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = project["name"].replace(" ", "_").replace("/", "_")
    filename = f"{safe_name}_manuals.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Bulk PDF download (ZIP) with progress tracking
# ---------------------------------------------------------------------------

def _build_download_zip(job_id: str, items: list[dict], project_name: str):
    """Background thread: build a ZIP of manual PDFs with per-item progress."""
    total = len(items)
    try:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for idx, item in enumerate(items, 1):
                brand = item.get("brand") or "unknown"
                model = item.get("model_number") or "unknown"

                _download_jobs[job_id] = {
                    "status": "building",
                    "message": f"Downloading {idx}/{total}: {brand} {model}",
                    "current": idx,
                    "total": total,
                }

                product = item.get("products")
                storage_path = product.get("manual_storage_path") if product else None
                pdf_bytes = None

                if storage_path:
                    try:
                        sb = db.get_client()
                        pdf_bytes = sb.storage.from_("manuals").download(storage_path)
                    except Exception:
                        pass

                if not pdf_bytes and item.get("manual_url"):
                    try:
                        resp = requests_lib.get(item["manual_url"], timeout=30)
                        if "pdf" in resp.headers.get("Content-Type", "").lower():
                            pdf_bytes = resp.content
                    except Exception:
                        pass

                if pdf_bytes:
                    safe_brand = brand.replace(" ", "_")
                    safe_model = model.replace(" ", "_")
                    zf.writestr(f"{safe_brand}_{safe_model}_manual.pdf", pdf_bytes)

        buffer.seek(0)
        safe_name = project_name.replace(" ", "_").replace("/", "_")

        _download_jobs[job_id] = {
            "status": "done",
            "message": "Download ready!",
            "current": total,
            "total": total,
            "file_bytes": buffer.getvalue(),
            "filename": f"{safe_name}_manuals.zip",
        }
    except Exception as e:
        logger.error("Download ZIP build failed: %s", e, exc_info=True)
        _download_jobs[job_id] = {
            "status": "error",
            "message": str(e),
            "current": 0,
            "total": total,
        }


@app.post("/api/projects/{project_id}/start-download")
async def start_download(project_id: str, data: dict = Body(default={}), user: dict = Depends(get_current_user)):
    """Start a background ZIP build and return a job ID for progress polling."""
    await _check_project_access(project_id, user)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    all_items = db.get_project_items(project_id)
    item_ids = data.get("ids") if data else None

    if item_ids:
        selected = set(item_ids)
        items = [i for i in all_items if i["id"] in selected]
    else:
        items = all_items

    items_with_manual = [i for i in items if i.get("manual_url")]
    if not items_with_manual:
        raise HTTPException(status_code=404, detail="No manuals available to download")

    job_id = str(uuid.uuid4())
    _download_jobs[job_id] = {
        "status": "building",
        "message": "Preparing download...",
        "current": 0,
        "total": len(items_with_manual),
    }

    thread = threading.Thread(
        target=_build_download_zip,
        args=(job_id, items_with_manual, project["name"]),
        daemon=True,
    )
    thread.start()

    return {"download_id": job_id, "total": len(items_with_manual)}


@app.get("/api/downloads/{job_id}/progress")
async def download_progress(job_id: str):
    """Poll download job progress."""
    job = _download_jobs.get(job_id)
    if not job:
        return {"status": "not_found", "message": "Download job not found"}
    return {k: v for k, v in job.items() if k not in ("file_bytes", "filename")}


@app.get("/api/downloads/{job_id}/file")
async def download_file(job_id: str):
    """Serve the completed ZIP file and clean up the job."""
    job = _download_jobs.get(job_id)
    if not job or job.get("status") != "done":
        raise HTTPException(status_code=404, detail="Download not ready")

    file_bytes = job.get("file_bytes")
    filename = job.get("filename", "manuals.zip")

    # Clean up
    del _download_jobs[job_id]

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Serve frontend (must be LAST — catches all unmatched routes)
# ---------------------------------------------------------------------------

_frontend_dir = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(_frontend_dir):
    app.mount("/", StaticFiles(directory=_frontend_dir, html=True), name="frontend")
